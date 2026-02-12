import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import numpy as np
from pythainlp.corpus import thai_stopwords
from pythainlp.tokenize import word_tokenize

from query_normalizer import QueryNormalizer


class QAPromptManager:
    """Manage QA prompt data from Excel and provide semantic + keyword matching."""

    COLUMN_SOURCE_DOCUMENT = "ภาคและเล่มเอกสาร"
    COLUMN_CATEGORY = "หมวดหมู่"
    COLUMN_QUESTION = "คำถาม"

    def __init__(
        self,
        excel_path: str = "./data/QA.xlsx",
        json_path: str = "./data/qa_prompt.json",
        embedding_model: Optional[Any] = None,
        match_threshold: float = 0.50,
        match_min_gap: float = 0.06,
        synonym_path: str = "./data/query_synonyms.json",
    ):
        self.excel_path = Path(excel_path)
        self.json_path = Path(json_path)
        self.embedding_model = embedding_model
        self.match_threshold = match_threshold
        self.match_min_gap = match_min_gap
        self.query_normalizer = QueryNormalizer(synonym_path=synonym_path)

        self._stopwords = set(thai_stopwords())
        self._items: List[Dict[str, Any]] = []
        self._answered_items: List[Dict[str, Any]] = []
        self._answered_embeddings: Optional[np.ndarray] = None
        self._answered_tokens: List[set[str]] = []
        self._answered_canonicals: List[set[str]] = []
        self._index_ready = False

    def sync_from_excel(self, force: bool = False, preserve_answers: bool = True) -> Dict[str, Any]:
        """Sync Excel -> JSON when needed, preserving non-empty answers from existing JSON."""
        should_sync = force or self._should_sync()
        if should_sync:
            items = self._build_items_from_excel(preserve_answers=preserve_answers)
            self._write_items(items)
        else:
            items = self.load_items()

        self._set_items(items)
        return {
            "updated": should_sync,
            "count": len(items),
            "excel_path": str(self.excel_path),
            "json_path": str(self.json_path),
        }

    def load_items(self) -> List[Dict[str, Any]]:
        """Load QA items from JSON."""
        if not self.json_path.exists():
            return []

        with self.json_path.open("r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, list):
            return [self._normalize_item(item) for item in data if isinstance(item, dict)]
        if isinstance(data, dict) and isinstance(data.get("items"), list):
            return [self._normalize_item(item) for item in data["items"] if isinstance(item, dict)]
        return []

    def find_best_match(self, query: str, expanded_query: Optional[str] = None) -> Optional[Dict[str, Any]]:
        """
        Find best answered QA candidate using hybrid score:
        combined_score = 0.7 * semantic + 0.3 * keyword
        """
        if not query or not query.strip():
            return None
        self._ensure_index()
        if not self._answered_items:
            return None
        if self.embedding_model is None:
            return None

        normalized_query_result = self.query_normalizer.normalize_query(query.strip())
        normalized_query = normalized_query_result["normalized_text"]

        variants: List[Dict[str, Any]] = []
        seen_variant_texts = set()

        def add_variant(raw_text: str) -> None:
            raw_text = raw_text.strip()
            if not raw_text:
                return
            variant_result = self.query_normalizer.normalize_query(raw_text)
            normalized_text = variant_result["normalized_text"]
            if not normalized_text or normalized_text in seen_variant_texts:
                return
            seen_variant_texts.add(normalized_text)
            variants.append(
                {
                    "raw_text": raw_text,
                    "normalized_text": normalized_text,
                    "matched_canonicals": variant_result["matched_canonicals"],
                }
            )

        add_variant(query)
        if normalized_query:
            add_variant(normalized_query)

        if expanded_query and expanded_query.strip():
            add_variant(expanded_query)

        if normalized_query_result["matched_canonicals"]:
            enriched_query = f"{normalized_query} {' '.join(normalized_query_result['matched_canonicals'])}".strip()
            add_variant(enriched_query)

        if not variants:
            return None

        variant_embeddings = self.embedding_model.encode(
            [variant["normalized_text"] for variant in variants],
            normalize_embeddings=True,
        )
        variant_tokens = [self._tokenize(variant["normalized_text"]) for variant in variants]
        variant_canonicals = [set(variant["matched_canonicals"]) for variant in variants]
        scored_candidates: List[Dict[str, Any]] = []

        for idx, item in enumerate(self._answered_items):
            item_best_score = -1.0
            item_best_semantic = 0.0
            item_best_keyword = 0.0
            item_best_variant = variants[0]["raw_text"]
            item_best_canonicals: List[str] = []

            for variant_idx, variant in enumerate(variants):
                semantic_score = float(np.dot(self._answered_embeddings[idx], variant_embeddings[variant_idx]))
                keyword_score = self._keyword_overlap(variant_tokens[variant_idx], self._answered_tokens[idx])
                canonical_score = self._canonical_overlap(
                    variant_canonicals[variant_idx],
                    self._answered_canonicals[idx],
                )
                combined_score = (0.7 * semantic_score) + (0.3 * keyword_score)
                combined_score = min(1.0, combined_score + (0.05 * canonical_score))

                if combined_score > item_best_score:
                    item_best_score = combined_score
                    item_best_semantic = semantic_score
                    item_best_keyword = keyword_score
                    item_best_variant = variant["raw_text"]
                    item_best_canonicals = variant["matched_canonicals"]

            candidate = dict(item)
            candidate["semantic_score"] = item_best_semantic
            candidate["keyword_score"] = item_best_keyword
            candidate["combined_score"] = item_best_score
            candidate["matched_variant"] = item_best_variant
            candidate["matched_canonicals"] = item_best_canonicals
            scored_candidates.append(candidate)

        if not scored_candidates:
            return None

        ranked_candidates = sorted(
            scored_candidates,
            key=lambda candidate: candidate["combined_score"],
            reverse=True,
        )
        top_candidate = ranked_candidates[0]
        second_score = ranked_candidates[1]["combined_score"] if len(ranked_candidates) > 1 else 0.0
        score_gap = top_candidate["combined_score"] - second_score

        if top_candidate["combined_score"] < self.match_threshold:
            return None
        if score_gap < self.match_min_gap:
            return None

        top_candidate["score_gap"] = score_gap
        top_candidate["normalized_query"] = normalized_query
        return top_candidate

    def _should_sync(self) -> bool:
        """Return True when JSON missing or Excel has newer mtime."""
        if not self.excel_path.exists():
            return False
        if not self.json_path.exists():
            return True
        return self.excel_path.stat().st_mtime > self.json_path.stat().st_mtime

    def _build_items_from_excel(self, preserve_answers: bool = True) -> List[Dict[str, Any]]:
        items = self._parse_excel_items()
        if not preserve_answers:
            return items

        existing_items = self.load_items()
        existing_by_id = {item["id"]: item for item in existing_items if item.get("id")}

        for item in items:
            old = existing_by_id.get(item["id"])
            if not old:
                continue
            old_answer = str(old.get("answer", "")).strip()
            if old_answer:
                item["answer"] = old.get("answer", "")
                item["answer_updated_at"] = old.get("answer_updated_at")

        return items

    def _parse_excel_items(self) -> List[Dict[str, Any]]:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required to read QA.xlsx. Please install with: pip install openpyxl"
            ) from exc

        workbook = load_workbook(filename=self.excel_path, data_only=True, read_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)

        headers = next(rows, None)
        if headers is None:
            workbook.close()
            return []

        header_index = {str(value).strip(): idx for idx, value in enumerate(headers) if value is not None}
        required_cols = [
            self.COLUMN_SOURCE_DOCUMENT,
            self.COLUMN_CATEGORY,
            self.COLUMN_QUESTION,
        ]

        missing = [col for col in required_cols if col not in header_index]
        if missing:
            workbook.close()
            raise ValueError(f"Missing required columns in Excel: {missing}")

        source_idx = header_index[self.COLUMN_SOURCE_DOCUMENT]
        category_idx = header_index[self.COLUMN_CATEGORY]
        question_idx = header_index[self.COLUMN_QUESTION]

        current_source = ""
        current_category = ""
        items: List[Dict[str, Any]] = []

        for row in rows:
            source_value = self._clean_cell(row[source_idx] if source_idx < len(row) else "")
            category_value = self._clean_cell(row[category_idx] if category_idx < len(row) else "")
            question_value = self._clean_cell(row[question_idx] if question_idx < len(row) else "")

            if source_value:
                current_source = source_value
            if category_value:
                current_category = category_value

            if not question_value:
                continue

            normalized_question = self._normalize_question(question_value)
            item_id = hashlib.sha1(normalized_question.encode("utf-8")).hexdigest()
            items.append(
                {
                    "id": item_id,
                    "source_document": current_source,
                    "category": current_category,
                    "question": question_value,
                    "answer": "",
                    "answer_updated_at": None,
                }
            )

        workbook.close()
        return items

    def _write_items(self, items: List[Dict[str, Any]]) -> None:
        self.json_path.parent.mkdir(parents=True, exist_ok=True)
        with self.json_path.open("w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)

    def _set_items(self, items: List[Dict[str, Any]]) -> None:
        self._items = items
        self._index_ready = False

    def _ensure_index(self) -> None:
        if self._index_ready:
            return
        if self.embedding_model is None:
            self._answered_items = []
            self._answered_embeddings = None
            self._answered_tokens = []
            self._answered_canonicals = []
            self._index_ready = True
            return
        if not self._items:
            self._items = self.load_items()

        self._answered_items = [item for item in self._items if str(item.get("answer", "")).strip()]
        if not self._answered_items:
            self._answered_embeddings = None
            self._answered_tokens = []
            self._answered_canonicals = []
            self._index_ready = True
            return

        normalized_questions: List[str] = []
        self._answered_tokens = []
        self._answered_canonicals = []
        for item in self._answered_items:
            normalized_result = self.query_normalizer.normalize_query(item["question"])
            normalized_question = normalized_result["normalized_text"] or self._normalize_question(item["question"])
            normalized_questions.append(normalized_question)
            self._answered_tokens.append(self._tokenize(normalized_question))
            self._answered_canonicals.append(set(normalized_result["matched_canonicals"]))

        self._answered_embeddings = self.embedding_model.encode(
            normalized_questions,
            normalize_embeddings=True,
        )
        self._index_ready = True

    def _tokenize(self, text: str) -> set[str]:
        words = word_tokenize(text, engine="newmm")
        filtered = set()
        for word in words:
            normalized = word.strip().lower()
            if len(normalized) < 2:
                continue
            if normalized in self._stopwords:
                continue
            if normalized.isdigit():
                continue
            filtered.add(normalized)
        return filtered

    @staticmethod
    def _keyword_overlap(query_tokens: set[str], question_tokens: set[str]) -> float:
        if not query_tokens:
            return 0.0
        return len(query_tokens.intersection(question_tokens)) / len(query_tokens)

    @staticmethod
    def _canonical_overlap(query_canonicals: set[str], question_canonicals: set[str]) -> float:
        if not query_canonicals:
            return 0.0
        return len(query_canonicals.intersection(question_canonicals)) / len(query_canonicals)

    @staticmethod
    def _clean_cell(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_question(question: str) -> str:
        cleaned = re.sub(r"\s+", " ", question.strip())
        return cleaned.lower()

    @staticmethod
    def _normalize_item(item: Dict[str, Any]) -> Dict[str, Any]:
        question = str(item.get("question", "")).strip()
        normalized_question = QAPromptManager._normalize_question(question)
        item_id = item.get("id") or hashlib.sha1(normalized_question.encode("utf-8")).hexdigest()
        return {
            "id": str(item_id),
            "source_document": str(item.get("source_document", "")).strip(),
            "category": str(item.get("category", "")).strip(),
            "question": question,
            "answer": item.get("answer", "") if item.get("answer") is not None else "",
            "answer_updated_at": item.get("answer_updated_at"),
        }
